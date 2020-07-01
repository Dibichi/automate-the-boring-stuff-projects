#! python3
#sheetInverter - inverts sheets aka rows become columns and vice versa

import openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

userDirectory = input('What directory is the workbook in?\n')
userInput = input('What workbook to choose?\n')
file = os.path.join(userDirectory, userInput)

wb = openpyxl.load_workbook(file)
sheet = wb.active

print("It's invert time")

rows = [] #creates a list for all rows
for row in range(1, sheet.max_row + 1):
    data = []
    for cell in range(1, sheet.max_column + 1): #goes through every cell
            cellValue = sheet.cell(row=row, column=cell).value
            data.append(cellValue) #the value of every cell in that row is appended to data
            sheet.cell(row=row, column=cell).value = ''
    rows.append(data) #each row has its own list within one big list of rows

#Basically this writes the sheet from up to down instead of left to right
columnValue = 1 #to start from the first column
for values in rows: #for each nested list
    rowValue = 1 #to start from first row every time to write the value
    for cell in values: #for the cells in the row
        sheet.cell(row = rowValue, column = columnValue).value = cell #writes it to that cell
        rowValue += 1 #goes down the rows until out of cells
    columnValue+=1  #moves over to the next column

wb.save(file)
print('done')
