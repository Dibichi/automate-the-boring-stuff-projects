#! python3
#blankRowInserter.py - inserts m amounts rows at row n
import sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

rowInsert = sys.argv[1]
numRows = sys.argv[2]
file = sys.argv[3]

wb = openpyxl.load_workbook(file)
sheet = wb.active

for rowNum in range(rowInsert, sheet.max_row):
    for column in range(1, sheet.max_column + 1):
        cell = get_column_letter(column) + str(rowNum)
        sheet[get_column_letter(column) + str(rowNum + numRows)] = sheet[cell].value

for row in range(rowInsert, rowInsert + numRows):
    for column in range(1, sheet.max_column + 1):
        sheet[get_column_letter(column) + str(row)] = ''

print('Executed')

wb.save(file)
