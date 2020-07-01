#! python3
#multiplicationTable.py - sys arg determines multiplcation table created
import os, openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
number = int(sys.argv[1]) + 2

wb = openpyxl.Workbook()
sheet = wb.active

boldFont = Font( bold = True)

sheet.title = 'Multiplication Table for ' + str(number)
for rowNum in range(2,number):
    for columnNum in range(2,number):
        columnLetter = get_column_letter(columnNum)
        sheet[columnLetter + str(rowNum)]  = (rowNum - 1) * (columnNum - 1)

for column in range(2, sheet.max_row + 1):
    cell = 'A' + str(column)
    sheet[cell] = column - 1
    sheet[cell].font = boldFont

for row in range(2,sheet.max_column + 1):
    cell = (get_column_letter(row)) + '1'
    sheet[cell] = row - 1
    sheet[cell].font = boldFont

wb.save('multiplicationTable.xlsx')
